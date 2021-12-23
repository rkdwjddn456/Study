#-*- coding: utf-8 -*-
import json
import os
import copy
import openpyxl

def main():
    ############ 공통 변수 ############
    dl_folder = "\\\\192.168.0.192/DL/"
    dc_folder = "\\\\192.168.0.192/DC/"
    #today = datetime.today().strftime("%Y-%m-%d")
    dl_wb = openpyxl.load_workbook('DL_Form.xlsx')
    dl_id_count = 0
    i = 9
    while True:
        if dl_wb['Rank_Total_Form']['A' + str(i)].value == "총계":
            break
        else:
            dl_id_count += 1
        i += 1

    dc_wb = openpyxl.load_workbook('DC_Form.xlsx')
    dc_id_count = 0
    i = 9
    while True:
        if dc_wb['Rank_Total_Form']['A' + str(i)].value == "총계":
            break
        else:
            dc_id_count += 1
        i += 1

    dl_wb.close()
    dc_wb.close()
    filtering_folder = ["delete","video","dldelete","dcteam00","save","_s_folder","one_folder"]

    ###########################################      DL 파트      #########################################################

    ############ DL 변수 ############
    dl_id_day_list = []
    dl_total_day_list = []
    dl_rank_day_list = []
    dl_class_day_list = []
    dl_rank_class_count_list = []
    dl_day_total_list = []
    day_list = []

    dl_road_class_list = [
    "flatness_A", "flatness_B", "flatness_C", "flatness_D", "flatness_E", "walkway_paved", "walkway_block", "paved_state_broken", "paved_state_normal",
    "block_state_broken", "block_state_normal", "block_kind_bad", "block_kind_good", "outcurb_rectangle", "outcurb_slide", "outcurb_rectangle_broken", "outcurb_slide_broken",
    "restspace", "sidegap_in", "sidegap_out", "sewer_cross", "sewer_line", "brailleblock_dot", "brailleblock_line", "brailleblock_dot_broken", "brailleblock_line_broken",
    "continuity_tree", "continuity_manhole", "ramp_yes", "ramp_no", "bicycleroad_broken", "bicycleroad_normal", "planecrosswalk_broken", "planecrosswalk_normal",
    "steepramp", "bump_slow", "bump_zigzag", "weed", "floor_normal", "floor_broken", "flowerbed", "parkspace", "tierbump", "stone", "enterrail", "fireshutter"
    ]
    dl_space_class_list = [
    "stair_normal", "stair_broken", "wall", "window_sliding", "window_casement", "pillar", "lift", "door_normal", "door_rotation", "lift_door",
    "resting_place_roof", "reception_desk", "protect_wall_protective", "protect_wall_guardrail", "protect_wall_kickplate", "handle_vertical", "handle_lever", "handle_circular",
    "lift_button_normal", "lift_button_openarea", "lift_button_layer", "lift_button_emergency", "direction_sign_left", "direction_sign_right", "direction_sign_straight", "direction_sign_exit",
    "sign_disabled_toilet", "sign_disabled_parking", "sign_disabled_elevator", "sign_disabled_ramp", "sign_disabled_callbell", "sign_disabled_icon", "braille_sign", "chair_multi",
    "chair_one", "chair_circular", "chair_back", "chair_handle", "number_ticket_machine", "beverage_vending_machine", "beverage_desk", "trash_can", "mailbox"
    ]

    dl_class_day_dict_form = {
    "day": "",
    "road":0,
    "space":0
    }
    dl_id_day_dict_form = {
    "id":"",
    "day":"",
    "flatness_A":0,
    "flatness_B":0,
    "flatness_C":0,
    "flatness_D":0,
    "flatness_E":0,
    "walkway_paved":0,
    "walkway_block":0,
    "paved_state_broken":0,
    "paved_state_normal":0,
    "block_state_broken":0,
    "block_state_normal":0,
    "block_kind_bad":0,
    "block_kind_good":0,
    "outcurb_rectangle":0,
    "outcurb_slide":0,
    "outcurb_rectangle_broken":0,
    "outcurb_slide_broken":0,
    "restspace":0,
    "sidegap_in" :0,
    "sidegap_out" :0,
    "sewer_cross" :0,
    "sewer_line" :0,
    "brailleblock_dot":0,
    "brailleblock_line":0,
    "brailleblock_dot_broken":0,
    "brailleblock_line_broken" :0,
    "continuity_tree":0,
    "continuity_manhole":0,
    "ramp_yes":0,
    "ramp_no":0,
    "bicycleroad_broken":0,
    "bicycleroad_normal":0,
    "planecrosswalk_broken":0,
    "planecrosswalk_normal" :0,
    "steepramp":0,
    "bump_slow":0,
    "bump_zigzag":0,
    "weed":0,
    "floor_normal":0,
    "floor_broken":0,
    "flowerbed":0,
    "parkspace":0,
    "tierbump":0,
    "stone":0,
    "enterrail":0,
    "fireshutter":0,

    "stair_normal":0,
    "stair_broken":0,
    "wall":0,
    "window_sliding":0,
    "window_casement":0,
    "pillar":0,
    "lift":0,
    "door_normal":0,
    "door_rotation":0,
    "lift_door":0,
    "resting_place_roof":0,
    "reception_desk":0,
    "protect_wall_protective":0,
    "protect_wall_guardrail":0,
    "protect_wall_kickplate":0,
    "handle_vertical":0,
    "handle_lever":0,
    "handle_circular":0,
    "lift_button_normal":0,
    "lift_button_openarea":0,
    "lift_button_layer":0,
    "lift_button_emergency":0,
    "direction_sign_left":0,
    "direction_sign_right":0,
    "direction_sign_straight":0,
    "direction_sign_exit":0,
    "sign_disabled_toilet":0,
    "sign_disabled_parking":0,
    "sign_disabled_elevator":0,
    "sign_disabled_ramp":0,
    "sign_disabled_callbell":0,
    "sign_disabled_icon":0,
    "braille_sign":0,
    "chair_multi":0,
    "chair_one":0,
    "chair_circular":0,
    "chair_back":0,
    "chair_handle":0,
    "number_ticket_machine":0,
    "beverage_vending_machine":0,
    "beverage_desk":0,
    "trash_can":0,
    "mailbox":0
    }
    dl_rank_day_dict_form = {
    "id":"",
    "day":"",
    "labeling_total_count":0,
    "labeling_class_count":0,
    "labeling_image_count":0,
    }
    dl_day_total_dict_form = {
    "day":"",
    "total_person":0,
    "labeling_total_count":0,
    "labeling_image_count":0,
    }

    ############ DC 변수 ############
    dc_id_day_list = []
    dc_id_day_dict_form = {
    "id":"",
    "day":"",
    "flatness_A":0,
    "flatness_B":0,
    "flatness_C":0,
    "flatness_D":0,
    "flatness_E":0,
    "walkway_paved":0,
    "walkway_block":0,
    "paved_state_broken":0,
    "paved_state_normal":0,
    "block_state_broken":0,
    "block_state_normal":0,
    "block_kind_bad":0,
    "block_kind_good":0,
    "outcurb_rectangle":0,
    "outcurb_slide":0,
    "outcurb_rectangle_broken":0,
    "outcurb_slide_broken":0,
    "restspace":0,
    "sidegap_in" :0,
    "sidegap_out" :0,
    "sewer_cross" :0,
    "sewer_line" :0,
    "brailleblock_dot":0,
    "brailleblock_line":0,
    "brailleblock_dot_broken":0,
    "brailleblock_line_broken" :0,
    "continuity_tree":0,
    "continuity_manhole":0,
    "ramp_yes":0,
    "ramp_no":0,
    "bicycleroad_broken":0,
    "bicycleroad_normal":0,
    "planecrosswalk_broken":0,
    "planecrosswalk_normal" :0,
    "steepramp":0,
    "bump_slow":0,
    "bump_zigzag":0,
    "weed":0,
    "floor_normal":0,
    "floor_broken":0,
    "flowerbed":0,
    "parkspace":0,
    "tierbump":0,
    "stone":0,
    "enterrail":0,
    "fireshutter":0,

    "stair_normal":0,
    "stair_broken":0,
    "wall":0,
    "window_sliding":0,
    "window_casement":0,
    "pillar":0,
    "lift":0,
    "door_normal":0,
    "door_rotation":0,
    "lift_door":0,
    "resting_place_roof":0,
    "reception_desk":0,
    "protect_wall_protective":0,
    "protect_wall_guardrail":0,
    "protect_wall_kickplate":0,
    "handle_vertical":0,
    "handle_lever":0,
    "handle_circular":0,
    "lift_button_normal":0,
    "lift_button_openarea":0,
    "lift_button_layer":0,
    "lift_button_emergency":0,
    "direction_sign_left":0,
    "direction_sign_right":0,
    "direction_sign_straight":0,
    "direction_sign_exit":0,
    "sign_disabled_toilet":0,
    "sign_disabled_parking":0,
    "sign_disabled_elevator":0,
    "sign_disabled_ramp":0,
    "sign_disabled_callbell":0,
    "sign_disabled_icon":0,
    "braille_sign":0,
    "chair_multi":0,
    "chair_one":0,
    "chair_circular":0,
    "chair_back":0,
    "chair_handle":0,
    "number_ticket_machine":0,
    "beverage_vending_machine":0,
    "beverage_desk":0,
    "trash_can":0,
    "mailbox":0
    }

    ###########################################      DL 통계 데이터 작성      #########################################################

    ###########################################      ID-날짜 시트      #########################################################
    for root, dirs, files in os.walk(dl_folder):
    # for root, dirs, files in os.walk("./DL"):
        # DL 폴더 필터링
        dirs[:] = [dir for dir in dirs if dir.lower() not in filtering_folder and dir[:2] != "DC"]

        if len(files) > 0:
            for file_name in files:
                check_in_list_flag = False
                idx = 0
                check_in_list_flag2 = False
                idx2 = 0

                if file_name.split(".")[-1] == "json":
                    if file_name.split(".")[0][-1].lower() == "s":
                        continue
                    print("DL : " + root + "/" + file_name)

                    dl_id = root.split("\\")[-2].split("_")[0]
                    # id = file_name.split("_")[0]
                    dl_day = root.split("\\")[-2].split("_")[1]

                    # json_file = open(root + "\\" + file_name, "rt", encoding="UTF8")
                    # jsonString = json.load(json_file)
                    # dl_day = jsonString.get("info")["date_created"]
                    # dl_day = dl_day[:4] + "-" + dl_day[5:7] + "-" + dl_day[8:10]
                    # dl_id = jsonString.get("annotations")[0].get("processor_id")

                    if file_name[0] == "G":
                        dc_id = "000002"
                        dc_day = "0000-00-00"
                    else :
                        dc_id = file_name.split("_")[0]
                        # id = root.split("\\")[-3]
                        dc_day = file_name.split("_")[1].split(" ")[0]

                    for i, dl_id_day in enumerate(dl_id_day_list):
                        if dl_id_day["id"] == dl_id and dl_id_day["day"] == dl_day:
                            idx = i
                            check_in_list_flag = True
                            break

                    for i, dc_id_day in enumerate(dc_id_day_list):
                        if dc_id_day["id"] == dc_id and dc_id_day["day"] == dc_day:
                            idx2 = i
                            check_in_list_flag2 = True
                            break

                    for i, dl_rank_day in enumerate(dl_rank_day_list):
                        if dl_rank_day["id"] == dl_id and dl_rank_day["day"] == dl_day:
                            dl_rank_day["labeling_image_count"] += 1
                            break

                    json_file = open(root + "\\" + file_name,"rt",encoding="UTF8")
                    jsonString = json.load(json_file)

                    if jsonString.get("annotations") != None:
                        category_id = jsonString.get("annotations")
                        category_key = "category_id"
                    else:
                        category_id = jsonString.get("shapes")
                        category_key = "label"

                    for shapes in category_id:
                        if check_in_list_flag:
                            dl_id_day_list[idx][shapes[category_key]] += 1
                        else:
                            dl_id_day_dict = copy.deepcopy(dl_id_day_dict_form)
                            dl_rank_day_dict = copy.deepcopy(dl_rank_day_dict_form)
                            dl_id_day_dict["id"] = dl_id
                            dl_rank_day_dict["id"] = dl_id
                            dl_id_day_dict["day"] = dl_day
                            dl_rank_day_dict["day"] = dl_day
                            dl_id_day_dict[shapes[category_key]] += 1
                            dl_rank_day_dict["labeling_image_count"] += 1
                            dl_id_day_list.append(dl_id_day_dict)
                            dl_rank_day_list.append(dl_rank_day_dict)
                            idx = len(dl_id_day_list) - 1
                            check_in_list_flag = True
                        if check_in_list_flag2:
                            dc_id_day_list[idx2][shapes[category_key]] += 1
                        else:
                            dc_id_day_dict = copy.deepcopy(dc_id_day_dict_form)
                            # dc_rank_day_dict = copy.deepcopy(dc_rank_day_dict_form)
                            dc_id_day_dict["id"] = dc_id
                            # dc_rank_day_dict["id"] = id
                            dc_id_day_dict["day"] = dc_day
                            # dc_rank_day_dict["day"] = day
                            dc_id_day_dict[shapes[category_key]] += 1
                            # dc_rank_day_dict["labeling_image_count"] += 1
                            dc_id_day_list.append(dc_id_day_dict)
                            # dc_rank_day_list.append(dc_rank_day_dict)
                            idx2 = len(dc_id_day_list) - 1
                            check_in_list_flag2 = True

    ###########################################      전체집계 - 날짜 시트      #########################################################
    for dl_id_day in dl_id_day_list:
        check_in_list_flag = False
        idx = 0

        day = dl_id_day["day"]

        for i, dl_total_day in enumerate(dl_total_day_list):
            if dl_total_day["day"] == day:
                idx = i
                check_in_list_flag = True
                break

        for key, value in dl_id_day.items():
            if check_in_list_flag:
                if key == "id" or key == "day":
                    continue
                dl_total_day_list[idx][key] += value
            else:
                dl_total_day_dict = copy.deepcopy(dl_id_day_dict_form)
                dl_total_day_dict["day"] = day
                dl_total_day_list.append(dl_total_day_dict)
                idx = len(dl_total_day_list) - 1
                check_in_list_flag = True

    ###########################################      랭킹-날짜 시트      #########################################################
    for dl_total_day in dl_total_day_list:
        dl_class_day_dict = copy.deepcopy(dl_class_day_dict_form)
        for key, value in dl_total_day.items():
            if key == "id":
                continue
            elif key == "day":
                dl_class_day_dict["day"] = value
            else:
                if key in dl_road_class_list:
                    dl_class_day_dict["road"] += value
                elif key in dl_space_class_list:
                    dl_class_day_dict["space"] += value
        dl_class_day_list.append(dl_class_day_dict)

    for dl_rank_day in dl_rank_day_list:
        if dl_rank_day["day"] not in day_list:
            day_list.append(dl_rank_day["day"])
        for dl_id_day in dl_id_day_list:
            if dl_id_day["id"] == dl_rank_day["id"] and dl_id_day["day"] == dl_rank_day["day"]:
                for key, value in dl_id_day.items():
                    if key == "id" or key == "day":
                        continue
                    dl_rank_day["labeling_total_count"] += value
                    if value != 0:
                        dl_rank_day["labeling_class_count"] += 1

    ###########################################      랭킹-누계 시트      #########################################################
    for dl_id_day in dl_id_day_list:
        check_in_list_flag = False
        idx = 0

        id = dl_id_day["id"]
        day = dl_id_day["day"]

        for i, dl_rank_class_count in enumerate(dl_rank_class_count_list):
            if dl_rank_class_count["id"] == id and dl_rank_class_count["day"] == day:
                idx = i
                check_in_list_flag = True
                break

        for key, value in dl_id_day.items():
            if check_in_list_flag:
                if key == "id" or key == "day":
                    continue
                elif value != 0:
                    dl_rank_class_count_list[idx]["labeling_class_count"] += 1
            else:
                dl_rank_class_count_dict = copy.deepcopy(dl_rank_day_dict_form)
                dl_rank_class_count_dict["id"] = id
                dl_rank_class_count_dict["day"] = day
                dl_rank_class_count_list.append(dl_rank_class_count_dict)
                idx = len(dl_day_total_list) - 1
                check_in_list_flag = True

    ###########################################      일자집계 시트      #########################################################
    for dl_rank_day in dl_rank_day_list:
        check_in_list_flag = False
        idx = 0

        day = dl_rank_day["day"]

        for i, dl_day_total in enumerate(dl_day_total_list):
            if dl_day_total["day"] == day:
                idx = i
                check_in_list_flag = True
                dl_day_total_list[idx]["total_person"] += 1
                break

        for key, value in dl_rank_day.items():
            if check_in_list_flag:
                if key == "labeling_total_count" or key == "labeling_image_count":
                    dl_day_total_list[idx][key] += value
                else :
                    continue
            else:
                dl_day_total_dict = copy.deepcopy(dl_day_total_dict_form)
                dl_day_total_dict["day"] = day
                dl_day_total_dict["total_person"] = 1
                dl_day_total_list.append(dl_day_total_dict)
                idx = len(dl_day_total_list) - 1
                check_in_list_flag = True

    ###########################################      엑셀 그리기      #########################################################
    day_list.sort()
    for day_idx in range(1,len(day_list) + 1):
        dl_wb = openpyxl.load_workbook('DL_Form.xlsx')
        statistics_day_list = day_list[0:day_idx]
        dl_day_total_sheet = dl_wb.copy_worksheet(dl_wb['Day_Total_Form'])

        ###########################################      일자집계 시트      #########################################################
        dl_day_total_sheet.title = "일자별 집계"
        dl_day_total_sheet['A1'] = "일자별 가공 실적 집계"
        dl_day_total_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        dl_class_total_dict = copy.deepcopy(dl_class_day_dict_form)
        for dl_class_day in dl_class_day_list:
            if dl_class_day["day"] not in statistics_day_list:
                continue
            else:
                dl_class_total_dict["road"] += dl_class_day["road"]
                dl_class_total_dict["space"] += dl_class_day["space"]

        dl_day_total_sheet['C4'] = dl_class_total_dict["road"]
        dl_day_total_sheet['C5'] = dl_class_total_dict["space"]
        dl_day_total_sheet['D4'] = dl_class_total_dict["road"] / dl_id_count
        dl_day_total_sheet['D5'] = dl_class_total_dict["space"] / dl_id_count
        dl_day_total_sheet['F8'] = len(statistics_day_list)

        for i, dl_day_total in enumerate(dl_day_total_list):
            if dl_day_total_list[i]["day"] not in statistics_day_list:
                break
            else:
                dl_day_total_sheet['A'+ str(i+10)] = i+1
                dl_day_total_sheet['B' + str(i + 10)] = dl_day_total_list[i]["day"]
                dl_day_total_sheet['C' + str(i + 10)] = dl_day_total_list[i]["total_person"]
                dl_day_total_sheet['D' + str(i + 10)] = dl_day_total_list[i]["labeling_total_count"]
                dl_day_total_sheet['E' + str(i + 10)] = dl_day_total_list[i]["labeling_image_count"]

        ###########################################      랭킹-누계 시트      #########################################################
        dl_rank_total_sheet = dl_wb.copy_worksheet(dl_wb['Rank_Total_Form'])
        dl_rank_total_sheet.title = '랭킹-누계'
        dl_rank_total_sheet['A1'] = "가공팀 전체 순위 (DL) - 누계"
        dl_rank_total_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        dl_rank_total_sheet['C4'] = dl_class_total_dict["road"]
        dl_rank_total_sheet['C5'] = dl_class_total_dict["space"]
        dl_rank_total_sheet['D4'] = dl_class_total_dict["road"] / dl_id_count
        dl_rank_total_sheet['D5'] = dl_class_total_dict["space"] / dl_id_count

        for dl_rank_day in dl_rank_day_list:
            if dl_rank_day["day"] not in statistics_day_list:
                continue
            i = 9
            while True:
                if dl_rank_total_sheet['A' + str(i)].value == "총계":
                    break
                if dl_rank_total_sheet['B' + str(i)].value == dl_rank_day["id"]:
                    dl_rank_total_sheet['E' + str(i)] = dl_rank_total_sheet['E' + str(i)].value + int(dl_rank_day["labeling_total_count"])
                    dl_rank_total_sheet['H' + str(i)] = dl_rank_total_sheet['H' + str(i)].value + int(dl_rank_day["labeling_image_count"])
                i += 1

        dl_id_total_list = []
        for dl_id_day in dl_id_day_list:
            if dl_id_day["day"] not in statistics_day_list:
                continue
            check_in_list_flag = False
            idx = 0

            id = dl_id_day["id"]

            for i, dl_id_total in enumerate(dl_id_total_list):
                if dl_id_total["id"] == id:
                    idx = i
                    check_in_list_flag = True
                    break

            for key, value in dl_id_day.items():
                if check_in_list_flag:
                    if key == "id" or key == "day":
                        continue
                    dl_id_total_list[idx][key] += value
                else:
                    dl_id_total_dict = copy.deepcopy(dl_id_day_dict_form)
                    dl_id_total_dict["id"] = id
                    dl_id_total_list.append(dl_id_total_dict)
                    idx = len(dl_id_total_list) - 1
                    check_in_list_flag = True

        for dl_id_total in dl_id_total_list:
            i = 9
            while True:
                if dl_rank_total_sheet['A' + str(i)].value == "총계":
                    break
                if dl_rank_total_sheet['B' + str(i)].value == dl_id_total["id"]:
                    sum = 0
                    for key, value in dl_id_total.items():
                        if key == "id" or key == "day":
                            continue
                        elif value != 0:
                            sum+=1
                    dl_rank_total_sheet['G' + str(i)] = sum
                i += 1

        ###########################################      랭킹-날짜 시트      #########################################################
        dl_rank_day_sheet = dl_wb.copy_worksheet(dl_wb['Rank_Day_Form'])
        dl_rank_day_sheet.title = '랭킹-' + statistics_day_list[-1]
        dl_rank_day_sheet['A1'] = "가공팀 전체 순위 (DL) - " + statistics_day_list[-1]
        dl_rank_day_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        dl_rank_day_sheet['A6'] = statistics_day_list[-1]
        dl_rank_day_sheet['C4'] = dl_class_total_dict["road"]
        dl_rank_day_sheet['C5'] = dl_class_total_dict["space"]
        dl_rank_day_sheet['D4'] = dl_class_total_dict["road"] / dl_id_count
        dl_rank_day_sheet['D5'] = dl_class_total_dict["space"] / dl_id_count

        for dl_class_day in dl_class_day_list:
            if dl_class_day["day"] == statistics_day_list[-1]:
                dl_rank_day_sheet['C6'] = dl_class_day["road"]
                dl_rank_day_sheet['C7'] = dl_class_day["space"]
                dl_rank_day_sheet['D6'] = dl_class_day["road"] / dl_id_count
                dl_rank_day_sheet['D7'] = dl_class_day["space"] / dl_id_count

        for dl_rank_day in dl_rank_day_list:
            if statistics_day_list[-1] == dl_rank_day["day"]:
                i = 11
                while True:
                    if dl_rank_day_sheet['A'+str(i)].value == "총계":
                        break
                    elif dl_rank_day_sheet['B'+str(i)].value == dl_rank_day["id"]:
                        dl_rank_day_sheet['E' + str(i)] = dl_rank_day["labeling_total_count"]
                        dl_rank_day_sheet['G' + str(i)] = dl_rank_day["labeling_class_count"]
                        dl_rank_day_sheet['H' + str(i)] = dl_rank_day["labeling_image_count"]
                    i += 1

        ###########################################      전체집계-누계 시트      #########################################################
        dl_total_total_sheet = dl_wb.copy_worksheet(dl_wb['Total_Total_Form'])
        dl_total_total_sheet.title = '전체집계-누계'
        dl_total_total_sheet['A1'] = "가공팀 전체 현황 (DL) - 누계"
        dl_total_total_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        for dl_total_day in dl_total_day_list:
            if dl_total_day["day"] not in statistics_day_list:
                continue
            idx = 5
            for key, value in dl_total_day.items():
                if key == "id" or key == "day":
                    continue
                if idx == 51:
                    idx = 54
                dl_total_total_sheet['I' + str(idx)] = dl_total_total_sheet['I' + str(idx)].value + int(value)
                idx += 1

        ###########################################      전체집계-날짜 시트      #########################################################
        dl_total_day_sheet = dl_wb.copy_worksheet(dl_wb['Total_Day_Form'])

        for dl_total_day in dl_total_day_list:
            if dl_total_day["day"] != statistics_day_list[-1]:
                continue
            else:
                idx = 4

                for key, value in dl_total_day.items():
                    if key == "id":
                        continue
                    if key == "day":
                        day = value
                        continue
                    if idx == 50:
                        idx = 53
                    dl_total_day_sheet['I' + str(idx)] = value
                    idx += 1

                dl_total_day_sheet.title = '전체집계-' + statistics_day_list[-1]
                dl_total_day_sheet['A1'] = "가공팀 전체 현황 (DL) - " + statistics_day_list[-1]

        ###########################################      ID-누계 시트      #########################################################
        for dl_id_total in dl_id_total_list:
            dl_id_total_sheet = dl_wb.copy_worksheet(dl_wb['Id_Total_Form'])
            idx = 5

            for key, value in dl_id_total.items():
                if key == "id":
                    id = value
                    continue
                elif key == "day":
                    continue
                if idx == 51:
                    idx = 54
                dl_id_total_sheet['J' + str(idx)] = value
                idx += 1

            dl_id_total_sheet.title = id + "-누계"
            dl_id_total_sheet['A1'] = id + " - 누계"
            dl_id_total_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        ###########################################      ID-날짜 시트      #########################################################
        for dl_id_day in dl_id_day_list:
            if dl_id_day["day"] != statistics_day_list[-1]:
                continue
            dl_id_day_sheet = dl_wb.copy_worksheet(dl_wb['Id_Day_Form'])
            idx = 4

            for key, value in dl_id_day.items():
                if key == "id":
                    id = value
                    continue
                elif key == "day":
                    day = value
                    continue
                if idx == 50:
                    idx = 53
                dl_id_day_sheet['J' + str(idx)] = value
                idx += 1

            dl_id_day_sheet.title = id + "-" + day
            dl_id_day_sheet['A1'] = id + "-" + day

        ###########################################      DL 엑셀 파일 저장      #########################################################
        dl_form_sheet_list = ['Day_Total_Form', 'Rank_Total_Form', 'Rank_Day_Form', 'Total_Total_Form', 'Total_Day_Form', 'Id_Total_Form', 'Id_Day_Form']
        for dl_form_sheet in dl_form_sheet_list:
            dl_wb.remove(dl_wb[dl_form_sheet])
        if not (os.path.isdir("./output")):
            os.makedirs(os.path.join("./output"))
        dl_wb.save('./output/DL_' + statistics_day_list[-1] + '.xlsx')

    ######################################################################################################################
    ######################################################################################################################
    ###########################################      DC 파트      #########################################################
    ######################################################################################################################
    ######################################################################################################################

    ############ DC 변수 ############
    dc_total_day_list = []
    dc_rank_day_list = []
    dc_class_day_list = []
    dc_class_id_day_list = []

    dc_road_class_list = [
    "flatness", "walkway", "paved_state", "block_state", "block_kind", "outcurb", "restspace", "sidegap", "sewer", "brailleblock", "continuity",
    "ramp", "bicycleroad", "planecrosswalk", "steepramp", "bump", "weed", "floor", "flowerbed", "parkspace", "tierbump", "stone", "enterrail", "fireshutter",
    ]
    dc_space_class_list = [
    "stair", "stair_broken", "wall", "window", "pillar", "lift", "door", "lift_door", "resting_place_roof", "reception_desk", "protect_wall", "handle",
    "lift_button", "direction_sign", "sign_disabled", "braille_sign", "chair", "chair_back", "chair_handle", "number_ticket_machine",
    "beverage_vending_machine", "beverage_desk", "trash_can", "mailbox"
    ]

    dc_class_day_dict_form = {
    "day": "",
    "road":0,
    "space":0
    }
    dc_rank_day_dict_form = {
    "id": "",
    "day": "",
    "checked_image_count": 0,
    "json_count": 0
    }
    dc_class_link_dict_form = {
    "flatness":["flatness_A","flatness_B","flatness_C","flatness_D","flatness_E"],
    "walkway":["walkway_paved","walkway_block"],
    "paved_state":["paved_state_broken","paved_state_normal"],
    "block_state":["block_state_broken","block_state_normal"],
    "block_kind":["block_kind_bad","block_kind_good"],
    "outcurb":["outcurb_rectangle","outcurb_slide","outcurb_rectangle_broken","outcurb_slide_broken"],
    "restspace":["restspace"],
    "sidegap":["sidegap_in","sidegap_out"],
    "sewer":["sewer_cross","sewer_line"],
    "brailleblock":["brailleblock_dot","brailleblock_line","brailleblock_dot_broken","brailleblock_line_broken"],
    "continuity":["continuity_tree","continuity_manhole"],
    "ramp":["ramp_yes","ramp_no"],
    "bicycleroad":["bicycleroad_broken","bicycleroad_normal"],
    "planecrosswalk":["planecrosswalk_broken","planecrosswalk_normal"],
    "steepramp":["steepramp"],
    "bump":["bump_slow","bump_zigzag"],
    "weed":["weed"],
    "floor":["floor_normal","floor_broken"],
    "flowerbed":["flowerbed"],
    "parkspace":["parkspace"],
    "tierbump":["tierbump"],
    "stone":["stone"],
    "enterrail":["enterrail"],
    "fireshutter":["fireshutter"],

    "stair":["stair_normal"],
    "stair_broken":["stair_broken"],
    "wall":["wall"],
    "window":["window_sliding","window_casement"],
    "pillar":["pillar"],
    "lift":["lift"],
    "door":["door_normal","door_rotation"],
    "lift_door":["lift_door"],
    "resting_place_roof":["resting_place_roof"],
    "reception_desk":["reception_desk"],
    "protect_wall":["protect_wall_protective","protect_wall_guardrail","protect_wall_kickplate"],
    "handle":["handle_vertical","handle_lever","handle_circular"],
    "lift_button":["lift_button_normal","lift_button_openarea","lift_button_layer","lift_button_emergency"],
    "direction_sign":["direction_sign_left","direction_sign_right","direction_sign_straight","direction_sign_exit"],
    "sign_disabled":["sign_disabled_toilet","sign_disabled_parking","sign_disabled_elevator","sign_disabled_ramp","sign_disabled_callbell","sign_disabled_icon"],
    "braille_sign":["braille_sign"],
    "chair":["chair_multi","chair_one","chair_circular"],
    "chair_back":["chair_back"],
    "chair_handle":["chair_handle"],
    "number_ticket_machine":["number_ticket_machine"],
    "beverage_vending_machine":["beverage_vending_machine"],
    "beverage_desk":["beverage_desk"],
    "trash_can":["trash_can"],
    "mailbox":["mailbox"],
    }
    dc_class_id_day_dict_form = {
    "id":"",
    "day":"",
    "flatness":0,
    "walkway":0,
    "paved_state":0,
    "block_state":0,
    "block_kind":0,
    "outcurb":0,
    "restspace":0,
    "sidegap":0,
    "sewer":0,
    "brailleblock":0,
    "continuity":0,
    "ramp":0,
    "bicycleroad":0,
    "planecrosswalk":0,
    "steepramp":0,
    "bump":0,
    "weed":0,
    "floor":0,
    "flowerbed":0,
    "parkspace":0,
    "tierbump":0,
    "stone":0,
    "enterrail":0,
    "fireshutter":0,

    "stair":0,
    "stair_broken":0,
    "wall":0,
    "window":0,
    "pillar":0,
    "lift":0,
    "door":0,
    "lift_door":0,
    "resting_place_roof":0,
    "reception_desk":0,
    "protect_wall":0,
    "handle":0,
    "lift_button":0,
    "direction_sign":0,
    "sign_disabled":0,
    "braille_sign":0,
    "chair":0,
    "chair_back":0,
    "chair_handle":0,
    "number_ticket_machine":0,
    "beverage_vending_machine":0,
    "beverage_desk":0,
    "trash_can":0,
    "mailbox":0
    }

    ###########################################      DC 통계 데이터 작성      #########################################################

    ###########################################      ID-날짜 시트      #########################################################
    for dc_id_day in dc_id_day_list:
        class_sum_by_id_dict = copy.deepcopy(dc_class_id_day_dict_form)
        for key, value in dc_id_day.items():
            if key == "id":
                class_sum_by_id_dict["id"] = value
                continue
            elif key == "day":
                class_sum_by_id_dict["day"] = value
                continue
            for class_link_key in dc_class_link_dict_form:
                if key in dc_class_link_dict_form[class_link_key]:
                    class_sum_by_id_dict[class_link_key] += value
        dc_class_id_day_list.append(class_sum_by_id_dict)

    ###########################################      전체집계 - 날짜 시트      #########################################################
    for dc_class_id_day in dc_class_id_day_list:
        check_in_list_flag = False
        idx = 0

        day = dc_class_id_day["day"]

        for i, dc_total_day in enumerate(dc_total_day_list):
            if dc_total_day["day"] == day:
                idx = i
                check_in_list_flag = True
                break

        for key, value in dc_class_id_day.items():
            if check_in_list_flag:
                if key == "id" or key == "day":
                    continue
                dc_total_day_list[idx][key] += value
            else:
                dc_total_day_dict = copy.deepcopy(dc_class_id_day_dict_form)
                dc_total_day_dict["day"] = day
                dc_total_day_list.append(dc_total_day_dict)
                idx = len(dc_total_day_list) - 1
                check_in_list_flag = True

    ###########################################      랭킹-날짜 시트      #########################################################
    for root, dirs, files in os.walk(dc_folder):
    # for root, dirs, files in os.walk("./DC/"):
        # DC 폴더 필터링
        dirs[:] = [dir for dir in dirs if dir.lower() not in filtering_folder]

        if len(files) > 0:
            for file_name in files:
                print("DC : " + root + "/" + file_name)
                check_in_list_flag = False
                idx = 0

                if file_name[0] == "G":
                    id = "000002"
                    day = "0000-00-00"
                elif file_name.split(".")[-1] in ["jpg","json"]:
                    id = file_name.split("_")[0]
                    day = file_name.split("_")[1].split(" ")[0]
                else:
                    continue

                for i, dc_rank_day in enumerate(dc_rank_day_list):
                    if dc_rank_day["id"] == id and dc_rank_day["day"] == day:
                        idx = i
                        check_in_list_flag = True
                        break

                if check_in_list_flag:
                    if file_name.split(".")[-1] == "jpg":
                        dc_rank_day_list[idx]["checked_image_count"] += 1
                    elif file_name.split(".")[-1] == "json":
                        if file_name.split(".")[0][-1].lower() == "s":
                            dc_rank_day_list[idx]["json_count"] += 1
                else:
                    dc_rank_day_dict = copy.deepcopy(dc_rank_day_dict_form)
                    dc_rank_day_dict["id"] = id
                    dc_rank_day_dict["day"] = day
                    if file_name.split(".")[-1] == "jpg":
                        dc_rank_day_dict["checked_image_count"] += 1
                    elif file_name.split(".")[-1] == "json":
                        if file_name.split(".")[0][-1].lower() == "s":
                            dc_rank_day_dict["json_count"] += 1
                    dc_rank_day_list.append(dc_rank_day_dict)

    for dc_total_day in dc_total_day_list:
        dc_class_day_dict = copy.deepcopy(dc_class_day_dict_form)
        for key, value in dc_total_day.items():
            if key == "id":
                continue
            elif key == "day":
                dc_class_day_dict["day"] = value
            else:
                if key in dc_road_class_list:
                    dc_class_day_dict["road"] += value
                elif key in dc_space_class_list:
                    dc_class_day_dict["space"] += value
        dc_class_day_list.append(dc_class_day_dict)

    day_list = []

    for dc_rank_day in dc_rank_day_list:
        if dc_rank_day["day"] not in day_list:
            day_list.append(dc_rank_day["day"])
    ###########################################      엑셀 그리기      #########################################################
    day_list.sort()
    for day_idx in range(1,len(day_list) + 1):
        dc_wb = openpyxl.load_workbook('DC_Form.xlsx')
        statistics_day_list = day_list[0:day_idx]

        ###########################################      랭킹-누계 시트      #########################################################
        dc_rank_total_sheet = dc_wb.copy_worksheet(dc_wb['Rank_Total_Form'])
        dc_rank_total_sheet.title = '랭킹-누계'
        dc_rank_total_sheet['A1'] = "수집팀 전체 순위 (DC)  - 누계"
        dc_rank_total_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        dc_class_total_dict = copy.deepcopy(dc_class_day_dict_form)

        for dc_class_day in dc_class_day_list:
            if dc_class_day["day"] not in statistics_day_list:
                continue
            dc_class_total_dict["road"] += dc_class_day["road"]
            dc_class_total_dict["space"] += dc_class_day["space"]

        dc_rank_total_sheet['C4'] = dc_class_total_dict["road"]
        dc_rank_total_sheet['C5'] = dc_class_total_dict["space"]
        dc_rank_total_sheet['D4'] = dc_class_total_dict["road"] / dc_id_count
        dc_rank_total_sheet['D5'] = dc_class_total_dict["space"] / dc_id_count

        for dc_rank_day in dc_rank_day_list:
            if dc_rank_day["day"] not in statistics_day_list:
                continue
            i = 9
            while True:
                if dc_rank_total_sheet['A' + str(i)].value == "총계":
                    break
                if dc_rank_total_sheet['B' + str(i)].value == dc_rank_day["id"]:
                    dc_rank_total_sheet['E' + str(i)] = dc_rank_total_sheet['E' + str(i)].value + int(dc_rank_day["checked_image_count"])
                    dc_rank_total_sheet['F' + str(i)] = dc_rank_total_sheet['F' + str(i)].value + int(dc_rank_day["json_count"])
                i += 1

        for dc_class_id_day in dc_class_id_day_list:
            if dc_class_id_day["day"] not in statistics_day_list:
                continue
            i = 9
            while True:
                if dc_rank_total_sheet['A' + str(i)].value == "총계":
                    break
                if dc_rank_total_sheet['B' + str(i)].value == dc_class_id_day["id"]:
                    for key, value in dc_class_id_day.items():
                        if key == "id" or key == "day":
                            continue
                        else:
                            dc_rank_total_sheet['H' + str(i)] = dc_rank_total_sheet['H' + str(i)].value + int(value)
                i += 1

        ###########################################      랭킹-날짜 시트      #########################################################
        dc_rank_day_sheet = dc_wb.copy_worksheet(dc_wb['Rank_Day_Form'])
        dc_rank_day_sheet.title = '랭킹-' + statistics_day_list[-1]
        dc_rank_day_sheet['A1'] = "수집팀 전체 순위 (DC) - " + statistics_day_list[-1]

        dc_rank_day_sheet['A5'] = statistics_day_list[-1]
        dc_rank_day_sheet['C3'] = dc_class_total_dict["road"]
        dc_rank_day_sheet['C4'] = dc_class_total_dict["space"]
        dc_rank_day_sheet['D3'] = dc_class_total_dict["road"] / dc_id_count
        dc_rank_day_sheet['D4'] = dc_class_total_dict["space"] / dc_id_count

        for dc_class_day in dc_class_day_list:
            if dc_class_day["day"] != statistics_day_list[-1]:
                continue
            else:
                dc_rank_day_sheet['C5'] = dc_class_day["road"]
                dc_rank_day_sheet['C6'] = dc_class_day["space"]
                dc_rank_day_sheet['D5'] = dc_class_day["road"] / dc_id_count
                dc_rank_day_sheet['D6'] = dc_class_day["space"] / dc_id_count

        for dc_rank_day in dc_rank_day_list:
            if dc_rank_day["day"] != statistics_day_list[-1]:
                continue
            else:
                i = 10
                while True:
                    if dc_rank_day_sheet['A' + str(i)].value == "총계":
                        break
                    elif dc_rank_day_sheet['B' + str(i)].value == dc_rank_day["id"]:
                        dc_rank_day_sheet['E' + str(i)] = dc_rank_day["checked_image_count"]
                        dc_rank_day_sheet['F' + str(i)] = dc_rank_day["json_count"]
                    i += 1

        for dc_class_id_day in dc_class_id_day_list:
            if dc_class_id_day["day"] != statistics_day_list[-1]:
                continue
            i = 9
            while True:
                if dc_rank_day_sheet['A' + str(i)].value == "총계":
                    break
                if dc_rank_day_sheet['B' + str(i)].value == dc_class_id_day["id"]:
                    for key, value in dc_class_id_day.items():
                        if key == "id" or key == "day":
                            continue
                        else:
                            dc_rank_day_sheet['H' + str(i)] = dc_rank_day_sheet['H' + str(i)].value + int(value)
                i += 1

        ###########################################      전체집계-누계 시트      #########################################################
        dc_total_total_sheet = dc_wb.copy_worksheet(dc_wb['Total_Form'])
        dc_total_total_sheet.title = '전체집계-누계'
        dc_total_total_sheet['A1'] = "수집팀 전체 현황 (DC) - " + statistics_day_list[-1]
        dc_total_total_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        dc_total_total_dict = copy.deepcopy(dc_class_id_day_dict_form)
        for dc_class_id_day in dc_class_id_day_list:
            if dc_class_id_day["day"] not in statistics_day_list:
                continue
            else:
                for key, value in dc_class_id_day.items():
                    if key == "id" or key == "day":
                        continue
                    dc_total_total_dict[key] += value

        idx = 5
        for key, value in dc_total_total_dict.items():
            if key == "id" or key == "day":
                continue
            if idx == 29:
                idx = 32
            dc_total_total_sheet['E' + str(idx)] = value
            idx += 1

        ###########################################      전체집계-날짜 시트      #########################################################
        dc_total_day_sheet = dc_wb.copy_worksheet(dc_wb['Total_Day_Form'])
        idx = 4

        for dc_total_day in dc_total_day_list:
            if dc_total_day["day"] != statistics_day_list[-1]:
                continue
            else:
                for key, value in dc_total_day.items():
                    if key == "id" or key == "day":
                        continue
                    if idx == 28:
                        idx = 31
                    dc_total_day_sheet['E' + str(idx)] = value
                    idx += 1

        dc_total_day_sheet.title = '전체집계-' + statistics_day_list[-1]
        dc_total_day_sheet['A1'] = "수집팀 전체 현황 (DC) - " + statistics_day_list[-1]

        ###########################################      ID-누계 시트      #########################################################
        dc_id_total_list = []
        for dc_class_id_day in dc_class_id_day_list:
            if dc_class_id_day["day"] not in statistics_day_list:
                continue
            check_in_list_flag = False
            idx = 0

            id = dc_class_id_day["id"]

            for i, dc_id_total in enumerate(dc_id_total_list):
                if dc_id_total["id"] == id:
                    idx = i
                    check_in_list_flag = True
                    break

            for key, value in dc_class_id_day.items():
                if check_in_list_flag:
                    if key == "id" or key == "day":
                        continue
                    dc_id_total_list[idx][key] += value
                else:
                    dc_id_total_dict = copy.deepcopy(dc_class_id_day_dict_form)
                    dc_id_total_dict["id"] = id
                    dc_id_total_list.append(dc_id_total_dict)
                    idx = len(dc_id_total_list) - 1
                    check_in_list_flag = True

        for dc_id_total in dc_id_total_list:
            dc_id_total_sheet = dc_wb.copy_worksheet(dc_wb['Total_Form'])
            idx = 5

            for key, value in dc_id_total.items():
                if key == "id":
                    id = value
                    continue
                elif key == "day":
                    continue
                if idx == 29:
                    idx = 32
                dc_id_total_sheet['E' + str(idx)] = value
                idx += 1

            dc_id_total_sheet.title = id + "-누계"
            dc_id_total_sheet['A1'] = id + " - 누계"
            dc_id_total_sheet['A2'] = statistics_day_list[0] + " ~ " + statistics_day_list[-1]

        ###########################################      ID-날짜 시트      #########################################################
        for dc_class_id_day in dc_class_id_day_list:
            if dc_class_id_day["day"] != statistics_day_list[-1]:
                continue
            dc_class_id_day_sheet = dc_wb.copy_worksheet(dc_wb['Total_Day_Form'])
            idx = 4

            for key, value in dc_class_id_day.items():
                if key == "id":
                    id = value
                    continue
                elif key == "day":
                    day = value
                    continue
                if idx == 28:
                    idx = 31
                dc_class_id_day_sheet['E' + str(idx)] = value
                idx += 1

            dc_class_id_day_sheet.title = id + "-" + day
            dc_class_id_day_sheet['A1'] = id + "-" + day
        ###########################################      DC 엑셀 파일 저장      #########################################################
        dc_form_sheet_list = ['Rank_Total_Form', 'Rank_Day_Form', 'Total_Form', 'Total_Day_Form']
        for dc_form_sheet in dc_form_sheet_list:
            dc_wb.remove(dc_wb[dc_form_sheet])
        if not (os.path.isdir("./output")):
            os.makedirs(os.path.join("./output"))
        dc_wb.save('./output/DC_' + statistics_day_list[-1] + '.xlsx')

    ###########################################      웹 파트      #########################################################
    # import jpype
    # import asposecells
    # jpype.startJVM()
    # from asposecells.api import *
    #
    # wb = Workbook("C:\\Users/gskim/PycharmProjects/statistics/test/DL.xlsx")
    # wb.calculateFormula();
    # wb.save("C:\\Users/gskim/PycharmProjects/statistics/test/DL.html")
    #
    # wb = Workbook("C:\\Users/gskim/PycharmProjects/statistics/test/DC.xlsx")
    # wb.calculateFormula();
    # wb.save("C:\\Users/gskim/PycharmProjects/statistics/test/DC.html")

try:
    main()
except Exception as e:
    print(e)
    while(True):
        continue